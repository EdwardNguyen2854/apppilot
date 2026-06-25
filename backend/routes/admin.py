"""
Admin API routes for AppPilot.

Provides endpoints for:
- Importing weekly ZIP files from users
- Merging data into master database
- Aggregated usage dashboard
- Exporting Excel/CSV reports
"""

import csv
import io
import json
import logging
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from fastapi import APIRouter, HTTPException, UploadFile, Query
from fastapi.responses import FileResponse

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/admin", tags=["admin"])


def init_admin_router(database, config):
    """Initialize router with dependencies."""

    @router.post("/import")
    async def import_weekly_zip(file: UploadFile) -> Dict:
        """
        Import weekly ZIP file from user submissions.

        Expected ZIP structure:
        - summary.json: metadata about the submission
        - sessions.csv: session records
        - events.csv: usage events
        - health_checks.csv: health check records
        - (optional) metrics.csv: process metrics
        """
        if not file.filename.endswith('.zip'):
            raise HTTPException(status_code=400, detail="File must be a ZIP archive")

        try:
            contents = await file.read()
            zip_buffer = io.BytesIO(contents)

            records_imported = 0
            records_skipped = 0
            submission_info = {}

            with zipfile.ZipFile(zip_buffer, 'r') as zf:
                # Read summary.json for metadata
                if 'summary.json' in zf.namelist():
                    summary_data = json.loads(zf.read('summary.json'))
                    submission_info = summary_data.get('submission', summary_data)
                else:
                    submission_info = {
                        'submitted_at': datetime.now().isoformat(),
                        'source': 'unknown'
                    }

                # Process sessions.csv
                if 'sessions.csv' in zf.namelist():
                    csv_content = zf.read('sessions.csv').decode('utf-8')
                    reader = csv.DictReader(io.StringIO(csv_content))
                    for row in reader:
                        success = database.import_weekly_session(row)
                        if success:
                            records_imported += 1
                        else:
                            records_skipped += 1

                # Process events.csv
                if 'events.csv' in zf.namelist():
                    csv_content = zf.read('events.csv').decode('utf-8')
                    reader = csv.DictReader(io.StringIO(csv_content))
                    for row in reader:
                        success = database.import_weekly_event(row)
                        if success:
                            records_imported += 1
                        else:
                            records_skipped += 1

                # Process health_checks.csv
                if 'health_checks.csv' in zf.namelist():
                    csv_content = zf.read('health_checks.csv').decode('utf-8')
                    reader = csv.DictReader(io.StringIO(csv_content))
                    for row in reader:
                        success = database.import_weekly_health_check(row)
                        if success:
                            records_imported += 1
                        else:
                            records_skipped += 1

                # Process metrics.csv (optional)
                if 'metrics.csv' in zf.namelist():
                    csv_content = zf.read('metrics.csv').decode('utf-8')
                    reader = csv.DictReader(io.StringIO(csv_content))
                    for row in reader:
                        success = database.import_weekly_metric(row)
                        if success:
                            records_imported += 1
                        else:
                            records_skipped += 1

            # Record the import
            week_id = submission_info.get('week_id', datetime.now().strftime('%Y-W%W'))
            machine_id = submission_info.get('machine_id', 'unknown')

            database.record_import(
                week_id=week_id,
                machine_id=machine_id,
                file_name=file.filename,
                records_imported=records_imported,
                records_skipped=records_skipped,
                status='completed'
            )

            logger.info(f"Imported {records_imported} records from {file.filename}, skipped {records_skipped}")

            return {
                "success": True,
                "file_name": file.filename,
                "week_id": week_id,
                "machine_id": machine_id,
                "records_imported": records_imported,
                "records_skipped": records_skipped,
                "submitted_at": submission_info.get('submitted_at', datetime.now().isoformat())
            }

        except zipfile.BadZipFile:
            raise HTTPException(status_code=400, detail="Invalid ZIP file")
        except Exception as e:
            logger.error(f"Error importing ZIP: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to import: {str(e)}")

    @router.get("/summary")
    async def get_admin_summary(
        app_id: Optional[str] = Query(None),
        user: Optional[str] = Query(None),
        week_id: Optional[str] = Query(None)
    ) -> Dict:
        """
        Get aggregated usage summary for admin dashboard.

        Returns: total sessions, total runtime, crashes, feature usage, etc.
        Supports filtering by app, user, week.
        """
        summary = database.get_admin_summary(app_id=app_id, user=user, week_id=week_id)
        return summary

    @router.post("/export")
    async def export_report(
        format: str = Query("csv", regex="^(csv|xlsx)$"),
        app_id: Optional[str] = Query(None),
        week_id: Optional[str] = Query(None)
    ) -> FileResponse:
        """
        Export aggregated data as CSV or Excel file.
        """
        # Get the data
        data = database.get_export_data(app_id=app_id, week_id=week_id)

        if format == 'csv':
            # Generate CSV
            output = io.StringIO()

            # Write sessions
            output.write("=== SESSIONS ===\n")
            if data.get('sessions'):
                writer = csv.DictWriter(output, fieldnames=data['sessions'][0].keys())
                writer.writeheader()
                writer.writerows(data['sessions'])

            output.write("\n=== EVENTS ===\n")
            if data.get('events'):
                writer = csv.DictWriter(output, fieldnames=data['events'][0].keys())
                writer.writeheader()
                writer.writerows(data['events'])

            output.write("\n=== HEALTH CHECKS ===\n")
            if data.get('health_checks'):
                writer = csv.DictWriter(output, fieldnames=data['health_checks'][0].keys())
                writer.writeheader()
                writer.writerows(data['health_checks'])

            csv_content = output.getvalue()
            output.close()

            # Write to temp file
            import tempfile
            temp_path = Path(tempfile.gettempdir()) / "opencode" / "export.csv"
            temp_path.parent.mkdir(parents=True, exist_ok=True)
            temp_path.write_text(csv_content)

            return FileResponse(
                path=str(temp_path),
                filename=f"apppilot_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                media_type="text/csv"
            )
        else:
            # Generate Excel
            try:
                from openpyxl import Workbook

                wb = Workbook()

                # Sessions sheet
                if data.get('sessions'):
                    ws = wb.active
                    ws.title = "Sessions"
                    ws.append(list(data['sessions'][0].keys()))
                    for row in data['sessions']:
                        ws.append(list(row.values()))

                # Events sheet
                if data.get('events'):
                    ws = wb.create_sheet("Events")
                    ws.append(list(data['events'][0].keys()))
                    for row in data['events']:
                        ws.append(list(row.values()))

                # Health checks sheet
                if data.get('health_checks'):
                    ws = wb.create_sheet("Health Checks")
                    ws.append(list(data['health_checks'][0].keys()))
                    for row in data['health_checks']:
                        ws.append(list(row.values()))

                import tempfile
                temp_path = Path(tempfile.gettempdir()) / "opencode" / "export.xlsx"
                temp_path.parent.mkdir(parents=True, exist_ok=True)
                wb.save(str(temp_path))

                return FileResponse(
                    path=str(temp_path),
                    filename=f"apppilot_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl not available")

    @router.get("/reports")
    async def get_reports(
        limit: int = Query(50, ge=1, le=200)
    ) -> List[Dict]:
        """List all imported reports with dates, record counts, status."""
        reports = database.get_import_history(limit=limit)
        return reports

    @router.get("/apps")
    async def get_all_app_stats(
        sort_by: str = Query("sessions", regex="^(sessions|runtime|crashes|app_name)$"),
        limit: int = Query(100, ge=1, le=500)
    ) -> List[Dict]:
        """
        Get usage stats per app.
        Returns total sessions, runtime, crashes sorted by usage.
        """
        stats = database.get_all_app_stats(sort_by=sort_by, limit=limit)
        return stats

    @router.get("/users")
    async def get_all_users(
        include_missing: bool = Query(False),
        limit: int = Query(100, ge=1, le=500)
    ) -> List[Dict]:
        """
        Get all users who submitted reports.
        Shows submission status, last seen, etc.
        """
        users = database.get_all_users(include_missing=include_missing, limit=limit)
        return users

    @router.get("/weekly-trend")
    async def get_weekly_trend(
        weeks: int = Query(12, ge=1, le=52)
    ) -> List[Dict]:
        """Get weekly usage trend data for the specified number of weeks."""
        trend = database.get_weekly_trend(weeks=weeks)
        return trend

    return router